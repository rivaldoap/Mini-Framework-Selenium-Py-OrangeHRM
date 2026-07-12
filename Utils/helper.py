import os
import time

from openpyxl import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select

### HIGHLIGHT ###
def Highlight(driver, element, blink=3):

    for i in range(blink):
        # HIGHLIGHT ON
        driver.execute_script("""
            arguments[0].style.border='2px solid fuchsia';
            arguments[0].style.boxShadow='0 0 5px fuchsia';
        """, element)
        time.sleep(0.2)

        # HIGHLIGHT OFF
        driver.execute_script("""
            arguments[0].style.border='';
            arguments[0].style.boxShadow='';
        """, element)
        time.sleep(0.2)

def navigate_to_menu(driver, excel_path):
    print(f" -> Membaca file dari path: {excel_path}")
    
    # Load workbook openpyxl
    wb = load_workbook(excel_path, data_only=True)
    sheet = wb.active
    
    # Melakukan iterasi baris, otomatis melompati baris ke-1 (Header)
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=5):
        
        # (A=RUN, C=MENU, D=SUBMENU)
        run_flag = row[0].value
        menu     = row[2].value
        submenu  = row[3].value
        # submenu2 = row[4].value
        
        # Skip jika kolom RUN diisi dengan "N" atau kosong
        if run_flag == "N" or run_flag is None:
            print(f"   [Global] Baris dilewati karena RUN = {run_flag}")
            continue
            
        # Navigasi Menu Utama secara dinamis
        if menu and str(menu).strip() != "":
            menu_clean = str(menu).strip()
            print(f"   [Global] Navigasi ke Menu: {menu_clean}")
            # Menggunakan string formatting f"" untuk memasukkan variabel menu ke dalam XPath
            xpath_menu = f"//a[.//span[text()='{menu_clean}']]"
            
            element = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, xpath_menu))
            )
            Highlight(driver, element)
            driver.find_element(By.XPATH, xpath_menu).click()
            time.sleep(2)
            
        # Navigasi Submenu1 secara dinamis
        if submenu and str(submenu).strip() != "":
            submenu_clean = str(submenu).strip()
            print(f"   [Global] Navigasi ke Submenu: {submenu_clean}")
            xpath_submenu = f"//a[normalize-space(text())='{submenu_clean}']"
            submenu_element = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, xpath_submenu))
            )
            Highlight(driver, submenu_element)
            driver.find_element(By.XPATH, xpath_submenu).click()
            time.sleep(2)

        # Navigasi Submenu2 secara dinamis
        # if submenu2 and str(submenu2).strip() != "":
        #     submenu2_clean = str(submenu2).strip()
        #     print(f"   [Global] Navigasi ke Sub-Submenu: {submenu2_clean}")
            
        #     # Catatan: Sesuaikan f-string XPath ini dengan struktur HTML OrangeHRM untuk sub-submenu kamu
        #     xpath_submenu2 = f"//a[normalize-space(text())='{submenu2_clean}']"
            
        #     submenu2_element = WebDriverWait(driver, 10).until(
        #         EC.element_to_be_clickable((By.XPATH, xpath_submenu2))
        #     )
        #     Highlight(driver, submenu2_element)
        #     driver.find_element(By.XPATH, xpath_submenu2).click()
        #     time.sleep(2)

# BASIC FUNCTION

def Exist(driver, locator, timeout=5):
    try:
        WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located(locator)
        )
        return True
    except:
        return False

def functionClick(driver, locator, timeout=5):
    try:
        # Tunggu sampai element clickable lalu klik biasa
        element = WebDriverWait(driver, timeout).until(
            EC.element_to_be_clickable(locator)
        )
        Highlight(driver, element)
        time.sleep(1.5)
        element.click()
    except Exception:
        # Jika terintersep animasi pop-up, pakai JS
        element = WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located(locator)
        )
        driver.execute_script("arguments[0].click();", element)

def functionInputText(driver, locator, text, timeout=10):
    element = WebDriverWait(driver, timeout).until(
        EC.visibility_of_element_located(locator)
    )
    Highlight(driver, element)
    time.sleep(1.5)
    if text is None:
        text = ""

    # Jika terintersep animasi pop-up, pakai JS
    driver.execute_script("arguments[0].click();", element)
    ### element.click()

    element.send_keys(Keys.CONTROL + "a")
    element.send_keys(Keys.BACKSPACE)
    element.send_keys(text)

#label_name: Nama label field input
#excel_value: Nilai yang ingin dipilih
def functionInputText_ByLabel(driver, label_name, excel_value, timeout=10):
    # get label name
    xpath_by_label = f"//div[contains(@class, 'oxd-input-group')][.//label[text()='{label_name}']]//*[self::input or self::textarea]"
    locator = (By.XPATH, xpath_by_label)
    
    element = WebDriverWait(driver, timeout).until(
        EC.visibility_of_element_located(locator)
    )
    
    Highlight(driver, element)
    time.sleep(1.5)
    
    if excel_value is None:
        excel_value = ""
    
    # Jika terintersep animasi pop-up, pakai JS
    driver.execute_script("arguments[0].click();", element)
    #### element.click()

    element.send_keys(Keys.CONTROL + "a")
    element.send_keys(Keys.BACKSPACE)
    element.send_keys(excel_value)

def functionClearInputText(driver, locator, timeout=10):
    element = WebDriverWait(driver, timeout).until(
        EC.visibility_of_element_located(locator)
    )
    
    Highlight(driver, element)
    element.click()
    element.send_keys(Keys.CONTROL + "a")
    element.send_keys(Keys.BACKSPACE)

#Select Radio Button
def functionSelectRDO(driver, locator, excel_value):
    xpath_string = locator[1] 
    
    # Masukkan nilai dari excel ke dalam '{}'
    xpath_dynamic = xpath_string.format(str(excel_value).strip())
    
    element = driver.find_element(By.XPATH, xpath_dynamic)
    
    # Jika terintersep animasi pop-up, pakai JS
    driver.execute_script("arguments[0].click();", element)

#label_name: Nama label di atas dropdown
#excel_value: Nilai yang ingin dipilih
def functionSelectDDL_ByLabel(driver, label_name, excel_value):
    if not excel_value:
        return # Skip jika di Excel kolom ini dikosongkan
        
    print(f"   [Dropdown] Mencoba memilih '{excel_value}' pada DDL '{label_name}'")
    
    # XPATH untuk mencari kotak dropdown berdasarkan nama Label di atasnya
    xpath_kotak_ddl = f"//div[div/label[text()='{label_name}']]//div[@class='oxd-select-text-input']"
    
    # Tunggu sampai kotak DDL bisa diklik, lalu klik untuk memunculkan opsi
    element = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, xpath_kotak_ddl))
    )
    Highlight(driver, element)
    element.click()
    time.sleep(1) # Berikan jeda animasi dropdown membuka
    
    # XPATH dinamis untuk menembak Opsi di dalam list yang teksnya sesuai Excel
    # Di OrangeHRM, list opsi menggunakan class 'oxd-select-option'
    xpath_all_in_one = (
        f"//div[@role='listbox']//div[@class='oxd-select-option' and .//span[contains(text(), '{excel_value}')]] | "
        f"//div[@role='listbox']//div[@class='oxd-select-option' and contains(normalize-space(), '{excel_value}')]"
    )
    
    try:
        # Tunggu sampai opsi yang dicari muncul di layar, lalu klik!
        opsi = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.XPATH, xpath_all_in_one))
        )
        opsi.click()
        time.sleep(1)

    except TimeoutException:
        # Jika gagal, tutup kembali kotak ddl agar tidak mengganggu elemen lain
        try: element.click()
        except: pass
        raise AssertionError(f"Gagal memilih dropdown '{label_name}': Opsi '{excel_value}' tidak ditemukan di layar.")

# SCROLL BERDASARKAN PIXEL (Misal: turun 500 pixel)
# Scroll layar berdasarkan koordinat X (horizontal) dan Y (vertical)
def functionScrollByPixel(driver, x=0, y=500):
    driver.execute_script(f"window.scrollBy({x}, {y});")
    time.sleep(1.5)


# Untuk scroll presisi ke object yang dituju
def functionScrollToElement(driver, locator, timeout=10):
    element = WebDriverWait(driver, timeout).until(
        EC.presence_of_element_located(locator)
    )
    # scroll element ke tengah layar
    driver.execute_script(
        "arguments[0].scrollIntoView({block: 'center'});", element
    )
    Highlight(driver, element)
    time.sleep(1.5)


# Scroll layar ke bagian paling atas halaman web
def functionScrollToTop(driver):
    driver.execute_script("window.scrollTo(0, 0);")
    time.sleep(1)


# Scroll layar keatas secara halus
def functionScrollToTopSmooth(driver):
    driver.execute_script("window.scrollTo({top: 0, behavior: 'smooth'});")
    time.sleep(1.5)

def verifyElementDisplayed(driver, locator, timeout=10):
    try:
        element = WebDriverWait(driver, timeout).until(
            EC.visibility_of_element_located(locator)
        )
        Highlight(driver, element)
        return True  # Berhasil ditemukan, kembalikan True
    except:
        return False # Tidak ditemukan, kembalikan False (Skrip TIDAK AKAN MATI)

def assertGetText(driver, locator, timeout=10):
    element = WebDriverWait(driver, timeout).until(
        EC.visibility_of_element_located(locator)
    )
    Highlight(driver, element)
    time.sleep(1.5)
    return element.text

# Memastikan suatu element muncul di layar halaman web
def assertElementDisplayed(driver, locator, timeout=10):
    try:
        element = WebDriverWait(driver, timeout).until(
            EC.visibility_of_element_located(locator)
        )
        Highlight(driver, element)
        # print(f"\n[ASSERT] Sukses! Element {locator} terlihat di layar.")
        assert element.is_displayed() is True
        print(f"[ASSERT] : '{locator}'")
    except:
        assert False, f"Gagal! Element dengan locator {locator} tidak ditemukan atau tidak muncul di layar."


# Assertion untuk memastikan text yang muncul di web sesuai dengan ekspektasi sesuai data dari excel
def assertTextEqualsValidasi(driver, locator, expected_text, timeout=10):
    try:
        # Tunggu sampai elemen benar-benar muncul di layar
        element = WebDriverWait(driver, timeout).until(
            EC.visibility_of_element_located(locator)
        )
        
        actual_text = element.text.strip()

        try:
            Highlight(driver, element)
        except:
            pass 
            
        print(f"[ASSERT] Menyamakan teks. Web: '{actual_text}' | Excel: '{expected_text}'")
    
        assert str(expected_text).strip() in actual_text, \
            f"Ekspektasi teks '{expected_text}' tidak cocok dengan teks di Web '{actual_text}'"
            
        print(f"[ASSERT] Sukses! Teks cocok sesuai ekspektasi.")
        return True

    except Exception as e:
        if isinstance(e, AssertionError):
            raise e
            
        # Jika terjadi Timeout/StaleElement, cetak log lalu paksa error dengan raise
        print(f"[ASSERT] Gagal! Elemen {locator} tidak ditemukan (stale/timeout).")
        raise AssertionError(f"Elemen {locator} tidak ditemukan atau tidak muncul di layar.")


def assertAndHandleBrowserAlert(driver, expected_text, timeout=10):
    print(f"   [Asersi] Memverifikasi apakah teks '{expected_text}' muncul di layar...")
    
    xpath_toast = f"//*[contains(normalize-space(text()), '{expected_text}')]"
    
    try:
        # Tunggu hingga elemen teks sukses/gagal dari Excel tersebut muncul di layar
        element = WebDriverWait(driver, timeout).until(
            EC.visibility_of_element_located((By.XPATH, xpath_toast))
        )
        Highlight(driver, element)
        print(f"   [Asersi] SUKSES: Teks '{expected_text}' ditemukan di layar!")
        
    except TimeoutException:
        # Jika tidak muncul, gagalkan tesnya dengan pesan yang jelas
        raise AssertionError(f"Gagal memverifikasi: Teks '{expected_text}' tidak muncul di layar setelah {timeout} detik.")

###Handle Browser Alert (Pop-up OK)
def functionHandleBrowser_AlertOK(driver, timeout=10):
    ###Fungsi global untuk menghandle JavaScript Browser Alert (Pop-up OK).
    ###Jika expected_alert_text diisi, fungsi akan memvalidasi teks di dalam alert tersebut.
    try:
        # Tunggu sampai Alert dari browser benar-benar muncul
        WebDriverWait(driver, timeout).until(EC.alert_is_present())
        
        # Pindah fokus Selenium ke jendela Alert tersebut
        alert = driver.switch_to.alert
        actual_text = alert.text.strip()
        print(f"[ALERT DETECTED] Teks pada alert: '{actual_text}'")
        
        # Klik tombol 'OK' pada alert
        alert.accept()
        print("[ALERT CLOSED] Berhasil klik OK pada browser alert.")
        
    except TimeoutException:
        raise AssertionError("Gagal! Browser alert tidak muncul setelah ditunggu.")
