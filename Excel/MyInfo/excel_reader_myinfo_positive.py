from openpyxl import load_workbook

FILE_PATH = r"Excel/MyInfo/My Info - Positive.xlsx"

def my_info_positive():
    SheetName = load_workbook(FILE_PATH)
    sheet = SheetName["MY_INFO"]

    all_data = []

    # Ambil baris pertama (Header) sebagai list untuk mencari indeks kolom nanti
    header_row = [cell.value for cell in sheet[1]]

    # Iterasi dari baris ke-2 (Data)
    for row in sheet.iter_rows(min_row=2, max_col=80, values_only=True):

        # row[0] adalah Kolom A (RUN)
        if row[0] is None:
            continue

        if str(row[0]).strip().upper() == "RUN":
            # KUNCI UTAMA: Fungsi pembantu (helper) untuk mengambil data berdasarkan nama kolom
            def get_val(column_name):
                if column_name in header_row:
                    idx = header_row.index(column_name)
                    return row[idx]
                return None
            
            # 3. Masukkan data menggunakan variabel nama kolom Excel
            all_data.append(
                {
                    "TC"                            : get_val("TC"),
                    "FIRST_NAME"                    : get_val("FIRST_NAME"),
                    "MID_NAME"                      : get_val("MID_NAME"),
                    "LAST_NAME"                     : get_val("LAST_NAME"),
                    "EMPLOYEE_ID"                   : get_val("EMPLOYEE_ID"),
                    "OTHER_ID"                      : get_val("OTHER_ID"),
                    "DRIVER_LICENSE_NUMBER"         : get_val("DRIVER_LICENSE_NUMBER"),
                    "LICENSE_EXP_DATE"              : get_val("LICENSE_EXP_DATE"),
                    "SSN_NUMBER"                    : get_val("SSN_NUMBER"),
                    "NATIONALITY"                   : get_val("NATIONALITY"),
                    "MARITAL_STATUS"                : get_val("MARITAL_STATUS"),
                    "DATE_OF_BIRTH"                 : get_val("DATE_OF_BIRTH"),
                    "GENDER"                        : get_val("GENDER"),
                    "BLOOD_TYPE"                    : get_val("BLOOD_TYPE"),
                    "TEST_FIELD"                    : get_val("TEST_FIELD"),
                    "STREET1"                       : get_val("STREET1"),
                    "STREET2"                       : get_val("STREET2"),
                    "STREET3"                       : get_val("STREET3"),
                    "CITY"                          : get_val("CITY"),
                    "STATE_PROVINCE"                : get_val("STATE_PROVINCE"),
                    "ZIP_POSTALCODE"                : get_val("ZIP_POSTALCODE"),
                    "COUNTRY"                       : get_val("COUNTRY"),
                    "TELP_HOME"                     : get_val("TELP_HOME"),
                    "TELP_MOBILE"                   : get_val("TELP_MOBILE"),
                    "TELP_WORK"                     : get_val("TELP_WORK"),
                    "WORK_EMAIL"                    : get_val("WORK_EMAIL"),
                    "OTHER_EMAIL"                   : get_val("OTHER_EMAIL"),
                    "NAME_EMERGENCYCONTACT"         : get_val("NAME_EMERGENCYCONTACT"),
                    "RELATIONSHIP_EMERGENCYCONTACT" : get_val("RELATIONSHIP_EMERGENCYCONTACT"),
                    "TELP_HOME_EMERGENCYCONTACT"    : get_val("TELP_HOME_EMERGENCYCONTACT"),
                    "MOBILE_EMERGENCYCONTACT"       : get_val("MOBILE_EMERGENCYCONTACT"), 
                    "TELP_WORK_EMERGENCYCONTACT"    : get_val("TELP_WORK_EMERGENCYCONTACT"),
                    "NAME_DEPENT"                   : get_val("NAME_DEPENT"),
                    "RELATIONSHIP_DEPENT"           : get_val("RELATIONSHIP_DEPENT"),
                    "PLEASE_SPECIFY_DEPENT"         : get_val("PLEASE_SPECIFY_DEPENT"),
                    "DATEOFBIRTH_DEPENT"            : get_val("DATEOFBIRTH_DEPENT"),
                    "DOC_PASSPORT_OR_VISA"          : get_val("DOC_PASSPORT_OR_VISA"),
                    "NUMBER"                        : get_val("NUMBER"),
                    "ISSUED_DATE"                   : get_val("ISSUED_DATE"),
                    "EXP_DATE"                      : get_val("EXP_DATE"),
                    "ELIGIBLE_STATUS"               : get_val("ELIGIBLE_STATUS"),
                    "ISSUED_BY"                     : get_val("ISSUED_BY"),
                    "ELIGIBLE_REVIEW_DATE"          : get_val("ELIGIBLE_REVIEW_DATE"),
                    "COMMENTS"                      : get_val("COMMENTS"),
                    "DECISION_WORKEXPERIENCE_Y_N"   : get_val("DECISION_WORKEXPERIENCE_Y_N"),
                    "WORKEXPRERIENCE_COMPANY"       : get_val("WORKEXPRERIENCE_COMPANY"),
                    "WORKEXPRERIENCE_JOB_TITLE"     : get_val("WORKEXPRERIENCE_JOB_TITLE"),
                    "WORKEXPRERIENCE_FROM"          : get_val("WORKEXPRERIENCE_FROM"),
                    "WORKEXPRERIENCE_TO"            : get_val("WORKEXPRERIENCE_TO"),
                    "WORKEXPRERIENCE_COMMENT"       : get_val("WORKEXPRERIENCE_COMMENT"),
                    "DECISION_EDUCATION_Y_N"        : get_val("DECISION_EDUCATION_Y_N"),
                    "EDUCATION_LEVEL"               : get_val("EDUATION_LEVEL"),
                    "EDUCATION_INSTITUTE"           : get_val("EDUCATION_INSTITUTE"),
                    "EDUCATION_MAJOR_SPECIALIZATION": get_val("EDUCATION_MAJOR_SPECIALIZATION"),
                    "EDUCATION_YEAR"                : get_val("EDUCATION_YEAR"),
                    "EDUCATION_GPA_SCORE"           : get_val("EDUCATION_GPA_SCORE"),
                    "EDUCATION_START_DATE"          : get_val("EDUCATION_START_DATE"),
                    "EDUCATION_END_DATE"            : get_val("EDUCATION_END_DATE"),
                    "DECISION_ADDSKILL_Y_N"         : get_val("DECISION_ADDSKILL_Y_N"),
                    "ADDSKILL_SKILL"                : get_val("ADDSKILL_SKILL"),
                    "ADDSKILL_YEARS_OF_EXPERIENCE"  : get_val("ADDSKILL_YEARS_OF_EXPERIENCE"),
                    "ADDSKILL_COMMENTS"             : get_val("ADDSKILL_COMMENTS"),
                    "DECISION_ADDLANGUAGES_Y_N"     : get_val("DECISION_ADDLANGUAGES_Y_N"),
                    "ADDLANGUAGE_LANGUAGE"          : get_val("ADDLANGUAGE_LANGUAGE"),
                    "ADDLANGUAGE_FLUENCY"           : get_val("ADDLANGUAGE_FLUENCY"),
                    "ADDLANGUAGE_COMPETENCY"        : get_val("ADDLANGUAGE_COMPETENCY"),
                    "ADDLANGUAGES_COMMENTS"         : get_val("ADDLANGUAGES_COMMENTS"),
                    "DECISION_ADDLICENSE_Y_N"       : get_val("DECISION_ADDLICENSE_Y_N"),
                    "ADDLICENSE_LICENSETYPE"        : get_val("ADDLICENSE_LICENSETYPE"),
                    "ADDLICENSE_LICENSENUMBER"      : get_val("ADDLICENSE_LICENSENUMBER"),
                    "ADDLICENSE_ISSUEDDATE"         : get_val("ADDLICENSE_ISSUEDDATE"),
                    "ADDLICENSE_EXPIRYDATE"         : get_val("ADDLICENSE_EXPIRYDATE"),
                    "MEMBERSHIP"                    : get_val("MEMBERSHIP"),
                    "SUBS_PAID_BY"                  : get_val("SUBS_PAID_BY"),
                    "SUBS_AMT"                      : get_val("SUBS_AMT"),
                    "CURRENCY"                      : get_val("CURRENCY"),
                    "SUBS_COMMENCEDATE"             : get_val("SUBS_COMMENCEDATE"),
                    "SUBS_RENEWALDATE"              : get_val("SUBS_RENEWALDATE"),
                }
            )

    return all_data