from openpyxl import load_workbook

FILE_PATH = r"Excel/Leave/Leave - Positive.xlsx"

def leave_positive():
    SheetName = load_workbook(FILE_PATH)
    sheet = SheetName["LEAVE"]

    all_data = []

    # min_row=2 : Otomatis skip baris 1 (Header)
    # max_col=5 : Membaca dari Kolom A (RUN) sampai max_col
    for row in sheet.iter_rows(min_row=2, max_col=15, values_only=True):

        # row[0] adalah Kolom A. Jika isinya "RUN", ambil data di baris tersebut
        if row[0] is None:
            continue

        if str(row[0]).strip().upper() == "RUN":
            all_data.append(
                {
                "TC"                        : row[1],
                "FROM_DATE"                 : row[4],
                "TO_DATE"                   : row[5],
                "SHOW_LEAVE_WITH_STATUS"    : row[6],
                "LEAVE_TYPE"                : row[7],
                "EMPLOYEE_NAME"             : row[8],
                "SUB_UNIT"                  : row[9],
                "INCLUDE_PAST_EMPLOYEES"    : row[10],
                "ASSERTION"                 : row[11],
                }
            )

    return all_data