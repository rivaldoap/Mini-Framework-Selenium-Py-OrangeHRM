from openpyxl import load_workbook

FILE_PATH = r"Excel/Admin/Admin - Positive.xlsx"

def admin_positive():
    SheetName = load_workbook(FILE_PATH)
    sheet = SheetName["ADD_USER"]

    all_data = []

    # min_row=2 : Otomatis skip baris 1 (Header)
    # max_col=5 : Membaca dari Kolom A (RUN) sampai max_col
    for row in sheet.iter_rows(min_row=2, max_col=15, values_only=True):

        # row[0] adalah Kolom A. Jika isinya "RUN", ambil data di baris tersebut
        if row[0] is None:
            continue

        if str(row[0]).strip().upper() == "RUN":
            all_data.append(
                {
                "TC"                : row[1],
                "USER_ROLE"         : row[4],
                "EMPLOYEE_NAME"     : row[5],
                "STATUS"            : row[6],
                "USERNAME"          : row[7],
                "PASSWORD"          : row[8],
                "CONFIRM_PASSWORD"  : row[9],
                "ASSERTION"         : row[10],
                }
            )

    return all_data