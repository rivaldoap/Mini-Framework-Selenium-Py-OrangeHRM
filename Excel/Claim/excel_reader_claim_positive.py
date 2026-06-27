from openpyxl import load_workbook

FILE_PATH = r"Excel/Claim/Claim - Positive.xlsx"

def claim_submitclaim_positive():
    SheetName = load_workbook(FILE_PATH)
    sheet = SheetName["SUBMIT_CLAIM"]

    all_data = []

    # min_row=2 : Otomatis skip baris 1 (Header)
    # max_col=5 : Membaca dari Kolom A (RUN) sampai max_col
    for row in sheet.iter_rows(min_row=2, max_col=15, values_only=True):

        # row[0] adalah Kolom A. Jika isinya "RUN", ambil data di baris tersebut
        if row[0] is None:
            continue

        if str(row[0]).strip().upper() == "RUN":
            all_data.append(
                {
                "TC"            : row[1],
                "SUBMENU"       : row[3],
                "EVENT"         : row[4],
                "CURRENCY"      : row[5],
                "REMARKS"       : row[6],
                "ASSERTION"     : row[7],
                }
            )

    return all_data


FILE_PATH = r"Excel/Claim/Claim - Positive.xlsx"

def claim_assign_positive():
    SheetName = load_workbook(FILE_PATH)
    sheet = SheetName["ASSIGN"]

    all_data = []

    # min_row=2 : Otomatis skip baris 1 (Header)
    # max_col=5 : Membaca dari Kolom A (RUN) sampai max_col
    for row in sheet.iter_rows(min_row=2, max_col=15, values_only=True):

        # row[0] adalah Kolom A. Jika isinya "RUN", ambil data di baris tersebut
        if row[0] is None:
            continue

        if str(row[0]).strip().upper() == "RUN":
            all_data.append(
                {
                "TC"            : row[1],
                "SUBMENU"       : row[3],
                "EMPLOYEE_NAME" : row[4],
                "EVENT"         : row[5],
                "CURRENCY"      : row[6],
                "REMARKS"       : row[7],
                "ASSERTION"     : row[8],
                }
            )

    return all_data