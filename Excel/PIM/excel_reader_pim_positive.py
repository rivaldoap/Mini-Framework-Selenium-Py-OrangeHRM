from openpyxl import load_workbook

FILE_PATH = r"Excel/PIM/PIM - Positive.xlsx"

def leave_positive():
    SheetName = load_workbook(FILE_PATH)
    sheet = SheetName["ADD_EMPLOYEE"]

    all_data = []

    # min_row=2 : Otomatis skip baris 1 (Header)
    # max_col=5 : Membaca dari Kolom A (RUN) sampai max_col
    for row in sheet.iter_rows(min_row=2, max_col=15, values_only=True):

        # row[0] adalah Kolom A. Jika isinya "RUN", ambil data di baris tersebut
        if row[0] is None:
            continue

        if str(row[0]).strip().upper() == "RUN":
            all_data.append(
                {
                "TC"                        : row[1],
                "FIRST_NAME"                : row[4],
                "MID_NAME"                  : row[5],
                "LAST_NAME"                 : row[6],
                "EMPLOYEE_ID"               : row[7],
                "CREATE_LOGIN_DETAILS"      : row[8],
                "USERNAME"                  : row[9],
                "STATUS"                    : row[10],
                "PASSWORD"                  : row[11],
                "CONFIRM_PASSWORD"          : row[12],
                }
            )

    return all_data